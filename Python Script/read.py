import sys
import os
from openpyxl.reader.excel import load_workbook

def getColumn(filepath,sheet,columnNum, rowNum):
    ''' return all the data in the columnNum 
    for rowNum from the filepath in the sheet'''
    wb = load_workbook(filepath)
    ws=wb[sheet]

    columnList = []
    data = []

    i = 2
    while(i in range(2,rowNum+1)):
        pn = str(ws.cell(row = i, column = columnNum).value)
        columnList.append(pn)
        i=i+1
    
    return print(columnList)

def getSheetsName(filepath):
    '''return all the sheet names 
    in the file saved in the filepath'''
    wb = load_workbook(filepath)
    return print(str(wb.get_sheet_names()))
    
def getSheetDim(filepath, sheet):
    '''return the sheet dimension row and column
    check until the cell is not equal to '' or None.
    in case the value it's zero the row or the column
    is number of the iteration in the while becasue
    the range has as highest limit get_highest_row/column
    so if this limit is the actually number of column or row
    the while return zero'''
    
    wb = load_workbook(filepath)
    ws=wb[sheet]
    
    numRow = 0
    numColumn = 0
    done=False
    j=1
    while(j in range(1,ws.get_highest_row()+1) and done == False):
        pn=ws.cell(row = j, column = 1).value 
        if pn != '' and pn != None:
            j=j+1
            done = False
        else:
            numRow = j
            done = True
    
    done=False
    i=1
    while(i in range(1,ws.get_highest_column())and done ==False):
        pn=ws.cell(row = 1, column = i).value 
        if pn != '' and pn != None:
            i=i+1
            done = False
        else:
            numColumn = i
            done = True
    
    if numColumn == 0:
        numColumn = i
    
    if numRow == 0:
        numRow = j
        
    return print(numRow-1,numColumn)

def allData(filePath, sheet, numRow, numColumn):
    '''return all the data in the sheet with numRow and numColumn as input'''
    wb = load_workbook(filePath)
    ws=wb[sheet]
    
    rowList = []
    data = []
    
    i=2
    
    for i in range(2,numRow+1):
        j=1
        for j in range(1,numColumn+1):
            cellValue = ws.cell(row=i,column=j).value
            rowList.append(cellValue)
    data.append(rowList)
      
    
    return print(data)

    
    
